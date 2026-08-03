#!/usr/bin/env python3
"""
列维-斯特劳斯神话学趋势预判 · 数据预处理脚本

处理大量消费者评论数据（≥500条），产出结构化初洗报告。
依赖：jieba（必需）；scikit-learn、openpyxl、xlrd（按输入与聚类需求可选）
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import jieba
import jieba.posseg as pseg

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

# ============================================================
# 情绪词典 — 扩展版（参考奥美情绪词 + 列维-斯特劳斯文化张力）
# ============================================================
# (情绪标签, 张力域, 权重)
EMOTION_LEXICON: dict[str, tuple[str, str, float]] = {
    # ---- 身份认同张力 ----
    "自卑": ("自卑/不自信", "身份认同", 0.9),
    "不自信": ("自卑/不自信", "身份认同", 0.9),
    "迷茫": ("迷茫/迷失", "身份认同", 0.9),
    "迷失": ("迷茫/迷失", "身份认同", 0.9),
    "真实": ("真实自我", "身份认同", 0.7),
    "做自己": ("真实自我", "身份认同", 0.9),
    "人设": ("人设/伪装", "身份认同", 0.8),
    "标签": ("标签化", "身份认同", 0.8),
    "定义": ("被定义", "身份认同", 0.7),
    "别人眼光": ("他人审视", "身份认同", 0.8),
    "面子": ("面子/体面", "身份认同", 0.7),
    "高级": ("高级感追求", "身份认同", 0.6),
    "精致": ("精致焦虑", "身份认同", 0.7),
    "土": ("土/low", "身份认同", 0.6),

    # ---- 社会角色张力 ----
    "压力": ("压力/重负", "社会角色", 0.8),
    "累": ("疲惫/透支", "社会角色", 0.7),
    "焦虑": ("焦虑/不安", "社会角色", 0.9),
    "不安": ("焦虑/不安", "社会角色", 0.8),
    "内卷": ("内卷/竞争", "社会角色", 0.9),
    "卷": ("内卷/竞争", "社会角色", 0.8),
    "平衡": ("角色平衡", "社会角色", 0.8),
    "职场": ("职场困境", "社会角色", 0.7),
    "装": ("装/表演", "社会角色", 0.8),
    "炫耀": ("炫耀/攀比", "社会角色", 0.8),
    "跟风": ("跟风/盲从", "社会角色", 0.7),
    "炒作": ("炒作/营销", "社会角色", 0.8),
    "网红": ("网红经济", "社会角色", 0.7),

    # ---- 情感缺失张力 ----
    "孤独": ("孤独/疏离", "情感缺失", 0.9),
    "寂寞": ("孤独/疏离", "情感缺失", 0.9),
    "冷漠": ("冷漠/疏远", "情感缺失", 0.8),
    "没人理解": ("不被理解", "情感缺失", 0.9),
    "陪伴": ("渴望陪伴", "情感缺失", 0.7),
    "爱": ("渴望爱", "情感缺失", 0.6),
    "共情": ("渴望共情", "情感缺失", 0.8),
    "共鸣": ("渴望共鸣", "情感缺失", 0.7),

    # ---- 生活方式张力 ----
    "自由": ("自由/解放", "生活方式", 0.8),
    "放松": ("放松/休息", "生活方式", 0.7),
    "自然": ("自然/本真", "生活方式", 0.7),
    "健康": ("健康/养生", "生活方式", 0.7),
    "慢": ("慢生活", "生活方式", 0.7),
    "简单": ("简单/极简", "生活方式", 0.6),
    "消费": ("消费主义", "生活方式", 0.7),
    "价格": ("价格敏感", "生活方式", 0.6),
    "性价比": ("性价比", "生活方式", 0.6),
    "智商税": ("智商税", "生活方式", 0.8),

    # ---- 文化二元对立特有 ----
    "异化": ("异化/变质", "文化张力", 0.9),
    "变味": ("异化/变质", "文化张力", 0.9),
    "泛滥": ("过度饱和", "文化张力", 0.8),
    "烂大街": ("过度饱和", "文化张力", 0.8),
    "赋予": ("赋魅/符号化", "文化张力", 0.7),
    "降级": ("降级/祛魅", "文化张力", 0.7),
    "本来": ("本质/原真", "文化张力", 0.6),
    "原来": ("本质/原真", "文化张力", 0.6),
    "以前": ("时间对比", "文化张力", 0.6),
    "现在": ("时间对比", "文化张力", 0.5),
    "我们那": ("地域对比", "文化张力", 0.7),
    "你们那": ("地域对比", "文化张力", 0.7),
    "上海": ("地域对比", "文化张力", 0.5),
    "本地": ("地域对比", "文化张力", 0.6),
    "外地": ("地域对比", "文化张力", 0.6),
}


# ============================================================
# 数据加载
# ============================================================

def detect_delimiter(path: Path) -> str:
    """自动检测 CSV/TSV 分隔符"""
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        first_line = f.readline()
    if "\t" in first_line:
        return "\t"
    return ","


def load_comments(path: str) -> list[dict[str, Any]]:
    """加载评论数据，统一返回 list[dict]"""
    p = Path(path)
    if not p.exists():
        print(f"❌ 文件不存在: {path}", file=sys.stderr)
        sys.exit(1)

    ext = p.suffix.lower()
    comments: list[dict[str, Any]] = []

    if ext == ".csv":
        delimiter = detect_delimiter(p)
        with open(p, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            for row in reader:
                comments.append(dict(row))

    elif ext in (".xlsx", ".xlsm"):
        if openpyxl is None:
            print("❌ 需要 openpyxl: pip install openpyxl", file=sys.stderr)
            sys.exit(1)
        wb = openpyxl.load_workbook(p, read_only=True)
        ws = wb.active
        if ws is None:
            print("❌ 工作簿无活动工作表", file=sys.stderr)
            sys.exit(1)
        header = [str(c.value or "") for c in next(ws.iter_rows(min_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(header):
                    row_dict[header[i]] = str(val or "")
            comments.append(row_dict)
        wb.close()

    elif ext in (".txt", ".md"):
        with open(p, "r", encoding="utf-8", errors="replace") as f:
            text = f.read()
        if ext == ".md":
            # 从 Markdown 中提取引用块和列表项
            lines = text.split("\n")
            for line in lines:
                stripped = line.strip()
                if stripped.startswith("> "):
                    comments.append({"comment": stripped[2:].strip()})
                elif stripped.startswith("- ") and len(stripped) > 4:
                    comments.append({"comment": stripped[2:].strip()})
        else:
            for line in text.split("\n"):
                stripped = line.strip()
                if stripped:
                    comments.append({"comment": stripped})

    else:
        print(f"❌ 不支持的格式: {ext}", file=sys.stderr)
        sys.exit(1)

    # 找到评论文本列
    comment_col = None
    col_names = list(comments[0].keys()) if comments else []
    for candidate in ["comment", "评论", "内容", "留言", "正文", "note-text", "text"]:
        if candidate in col_names:
            comment_col = candidate
            break
    if not comment_col and col_names:
        comment_col = col_names[0]  # 默认第一列

    # 找到点赞列
    like_col = None
    for candidate in ["count", "count 2", "点赞", "赞", "like", "likes", "互动"]:
        if candidate in col_names:
            like_col = candidate
            break

    # 结构化输出
    result = []
    for row_number, row in enumerate(comments, start=2):
        if not comment_col:
            continue
        text_val = str(row.get(comment_col, "")).strip()
        if len(text_val) < 3:
            continue  # 过滤太短发帖
        like_val = 0
        if like_col:
            try:
                like_val = int(float(str(row.get(like_col, 0)).replace(",", "").replace("赞", "0")))
            except (ValueError, TypeError):
                like_val = 0
        result.append({
            "source_id": f"S-{row_number:06d}",
            "text": text_val,
            "likes": like_val,
            "location": str(row.get("location", row.get("地点", ""))) if "location" in row or "地点" in row else "",
        })

    return result


# ============================================================
# 分词与词频
# ============================================================

def segment_and_freq(comments: list[dict[str, Any]]) -> dict[str, Any]:
    """jieba 分词 + 词性标注 + 词频统计（按评论数，非按词频）"""
    texts = [c["text"] for c in comments]
    likes = [c["likes"] for c in comments]

    # 词频统计：按"出现该词的评论数"统计
    word_comment_count: Counter = Counter()
    # 词性汇总
    pos_tags: Counter = Counter()
    # 加权词频（按点赞数加权）
    word_weighted: defaultdict[str, float] = defaultdict(float)
    # 每个词的最佳例句
    word_examples: dict[str, list[tuple[str, int]]] = defaultdict(list)

    for text, like in zip(texts, likes):
        # 分词 + 词性标注
        words = pseg.cut(text)
        seen_in_this_comment = set()
        for word, flag in words:
            w = word.strip()
            if len(w) < 2:
                continue
            # 过滤停用词（标点、空格、纯数字）
            if w in ("的", "了", "是", "在", "我", "有", "和", "就", "不", "人", "都",
                     "一", "一个", "也", "这", "他", "她", "它", "我们", "他们", "你们",
                     "什么", "怎么", "为什么", "这个", "那个", "没有", "不是", "可以",
                     "会", "要", "能", "让", "把", "被", "还", "又", "但", "而", "或",
                     "因为", "所以", "如果", "虽然", "然后", "之后", "之前", "时候",
                     "已经", "可能", "比较", "还是", "就是", "真的", "其实"):
                continue
            if w not in seen_in_this_comment:
                word_comment_count[w] += 1
                seen_in_this_comment.add(w)
            pos_tags[flag] += 1
            word_weighted[w] += like + 1  # +1 避免0权重

            # 保存前5条最高赞例句
            ex_list = word_examples[w]
            ex_list.append((text, like))
            ex_list.sort(key=lambda x: -x[1])
            if len(ex_list) > 5:
                ex_list.pop()

    # 排序
    top_words_all = word_comment_count.most_common(50)
    top_nouns = []
    top_adjs = []
    top_verbs = []

    for w, cnt in word_comment_count.most_common(200):
        # 再分词一次确认词性（缓存不够好，简单处理）
        words_p = list(pseg.cut(w))
        if words_p:
            flag = words_p[0].flag
            if flag.startswith("n") and cnt >= 3:
                top_nouns.append((w, cnt))
            elif flag.startswith("a") and cnt >= 3:
                top_adjs.append((w, cnt))
            elif flag.startswith("v") and cnt >= 3:
                top_verbs.append((w, cnt))

    return {
        "total_comments": len(comments),
        "total_likes": sum(likes),
        "top_words": [(w, c, word_weighted.get(w, 0)) for w, c in top_words_all[:25]],
        "top_nouns": top_nouns[:15],
        "top_adjs": top_adjs[:15],
        "top_verbs": top_verbs[:15],
        "word_examples": dict(word_examples),
        "pos_tags": dict(pos_tags.most_common(10)),
    }


# ============================================================
# 情绪扫描
# ============================================================

def emotion_scan(comments: list[dict[str, Any]]) -> dict[str, Any]:
    """扫描评论中的情绪词，按张力域汇总"""
    domain_scores: defaultdict[str, float] = defaultdict(float)
    domain_comment_counts: defaultdict[str, int] = defaultdict(int)
    emotion_comment_counts: Counter = Counter()
    emotion_examples: defaultdict[str, list[tuple[str, int]]] = defaultdict(list)

    for c in comments:
        text = c["text"]
        like = c["likes"]
        matched_domains = set()
        matched_emotions = set()

        for keyword, (emotion, domain, weight) in EMOTION_LEXICON.items():
            if keyword in text:
                if domain not in matched_domains:
                    domain_scores[domain] += weight * (like + 1)
                    domain_comment_counts[domain] += 1
                    matched_domains.add(domain)
                if emotion not in matched_emotions:
                    emotion_comment_counts[emotion] += 1
                    matched_emotions.add(emotion)
                    ex_list = emotion_examples[emotion]
                    ex_list.append((text, like))
                    ex_list.sort(key=lambda x: -x[1])
                    if len(ex_list) > 5:
                        ex_list.pop()

    return {
        "domain_scores": dict(sorted(domain_scores.items(), key=lambda x: -x[1])),
        "domain_comment_counts": dict(domain_comment_counts),
        "emotion_comment_counts": dict(emotion_comment_counts.most_common(20)),
        "emotion_examples": dict(emotion_examples),
    }


# ============================================================
# 简单聚类（基于关键词共现）
# ============================================================

def keyword_clustering(comments: list[dict[str, Any]], num_clusters: int = 5) -> list[dict[str, Any]]:
    """
    基于关键词语义的 KMeans 聚类。
    每个评论用高频词向量表示，用 sklearn KMeans 聚类。
    回退策略：如果 sklearn 不可用，用关键词规则分类。
    """
    texts = [c["text"] for c in comments]
    num_comments = len(texts)
    if num_comments == 0:
        return []

    # 过滤太短文本，用 TF-IDF
    processed = []
    valid_indices = []
    for i, t in enumerate(texts):
        words = " ".join(jieba.lcut(t))
        if len(words.strip()) > 2:
            processed.append(words)
            valid_indices.append(i)

    if len(processed) < num_clusters:
        # 文本太少，退回到单聚类
        return [{
            "cluster_id": 0,
            "method": "single_cluster_small_sample",
            "size": len(comments),
            "percentage": 100.0,
            "top_keywords": [],
            "representative_quotes": [
                {"source_id": c["source_id"], "text": c["text"][:120], "likes": c["likes"]}
                for c in sorted(comments, key=lambda x: -x["likes"])[:5]
            ],
            "member_source_ids": [c["source_id"] for c in comments],
            "emotional_tone": "混合",
        }]

    try:
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.cluster import KMeans

        vectorizer = TfidfVectorizer(max_features=500, min_df=2)
        X = vectorizer.fit_transform(processed)
        feature_names = vectorizer.get_feature_names_out()

        k = min(num_clusters, len(processed) // 5)  # ensure enough samples per cluster
        k = max(k, 2)
        km = KMeans(n_clusters=k, random_state=42, n_init="auto")
        labels = km.fit_predict(X)

        # 构建每个聚类的结果
        cluster_map: dict[int, list[int]] = {}
        for label_idx, label in enumerate(labels):
            original_idx = valid_indices[label_idx]
            cluster_map.setdefault(int(label), []).append(original_idx)

        result = []
        for cluster_id, indices in cluster_map.items():
            cluster_texts = [texts[i] for i in indices]
            cluster_comments = [comments[i] for i in indices]

            # 聚类内高频词
            cluster_words: Counter = Counter()
            for t in cluster_texts:
                for w in jieba.lcut(t):
                    if len(w) >= 2:
                        cluster_words[w] += 1

            # 按点赞排序
            sorted_c = sorted(cluster_comments, key=lambda x: -x["likes"])

            # TF-IDF 关键词（该聚类的 top 特征）
            centroid = km.cluster_centers_[cluster_id]
            top_feature_indices = centroid.argsort()[-8:][::-1]
            cluster_keywords = []
            for fi in top_feature_indices:
                if fi < len(feature_names):
                    cluster_keywords.append(feature_names[fi])

            result.append({
                "cluster_id": cluster_id,
                "method": "jieba_tfidf_kmeans",
                "size": len(indices),
                "percentage": round(len(indices) / num_comments * 100, 1),
                "top_keywords": cluster_keywords or [w for w, _ in cluster_words.most_common(8)],
                "representative_quotes": [
                    {"source_id": c["source_id"], "text": c["text"][:120], "likes": c["likes"]}
                    for c in sorted_c[:5]
                ],
                "member_source_ids": [c["source_id"] for c in cluster_comments],
                "emotional_tone": "待分析",
            })

        result.sort(key=lambda x: -x["size"])
        return result

    except Exception as e:
        # 通用安全回退：不使用任何业务领域关键词，不伪造多个主题。
        print(f"  ⚠ KMeans 不可用 ({e})，输出未聚类阅读导航", file=sys.stderr)
        global_words: Counter = Counter()
        for text in texts:
            for word in jieba.lcut(text):
                if len(word.strip()) >= 2:
                    global_words[word.strip()] += 1
        sorted_comments = sorted(comments, key=lambda item: -item["likes"])
        return [{
            "cluster_id": 0,
            "method": "unclustered_dependency_fallback",
            "size": len(comments),
            "percentage": 100.0,
            "top_keywords": [word for word, _ in global_words.most_common(12)],
            "representative_quotes": [
                {"source_id": item["source_id"], "text": item["text"][:120], "likes": item["likes"]}
                for item in sorted_comments[:8]
            ],
            "member_source_ids": [c["source_id"] for c in comments],
            "emotional_tone": "未聚类—待模型阅读",
        }]


# ============================================================
# 主流程
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="列维-斯特劳斯神话学趋势预判 · 数据预处理"
    )
    parser.add_argument("input_file", help="输入文件路径（CSV/TXT/MD）")
    parser.add_argument("output_dir", help="输出目录")
    parser.add_argument("--source", default="未知来源", help="数据来源平台名称")
    parser.add_argument("--cluster", type=int, default=5, help="聚类数（默认5）")
    args = parser.parse_args()

    input_path = Path(args.input_file)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"📂 读取数据: {input_path}")
    comments = load_comments(str(input_path))
    total = len(comments)

    if total < 10:
        print("❌ 有效评论不足10条，无法分析", file=sys.stderr)
        sys.exit(1)

    print(f"✅ 读取 {total} 条评论")

    # ---- 1. 分词与词频 ----
    print("📊 分词统计中...")
    freq_result = segment_and_freq(comments)

    # ---- 2. 情绪扫描 ----
    print("🔍 情绪扫描中...")
    emotion_result = emotion_scan(comments)

    # ---- 3. 聚类分析 ----
    print(f"📦 聚类分析（{args.cluster}类）...")
    clusters = keyword_clustering(comments, num_clusters=args.cluster)

    # ---- 4. 生成初洗报告 ----
    report_path = output_dir / "初洗报告.md"
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    with open(report_path, "w", encoding="utf-8") as f:
        f.write(f"# 初洗报告\n\n")
        f.write(f"- **数据来源**: {args.source}\n")
        f.write(f"- **总评论数**: {total}\n")
        f.write(f"- **总点赞/互动**: {freq_result['total_likes']:,}\n")
        f.write(f"- **生成时间**: {now}\n\n")
        f.write("---\n\n")

        # 高频词
        f.write("## 高频词 Top 25\n\n")
        f.write("| 词 | 提及评论数 | 加权分 | 例句 |\n")
        f.write("|---|---|---|---|\n")
        for word, count, weight in freq_result["top_words"]:
            examples = freq_result["word_examples"].get(word, [])
            example_text = examples[0][0][:60] + "…" if examples else ""
            f.write(f"| {word} | {count} | {weight:.0f} | {example_text} |\n")

        f.write("\n### 名词 Top 15\n")
        for w, c in freq_result.get("top_nouns", []):
            f.write(f"- {w} ({c})\n")

        f.write("\n### 形容词 Top 15\n")
        for w, c in freq_result.get("top_adjs", []):
            f.write(f"- {w} ({c})\n")

        f.write("\n### 动词 Top 15\n")
        for w, c in freq_result.get("top_verbs", []):
            f.write(f"- {w} ({c})\n")

        f.write("\n---\n\n")

        # 情绪矩阵
        f.write("## 情绪矩阵\n\n")
        f.write("### 张力域分布\n\n")
        f.write("| 张力域 | 加权分 | 提及评论数 |\n")
        f.write("|---|---|---|\n")
        for domain, score in emotion_result["domain_scores"].items():
            cnt = emotion_result["domain_comment_counts"].get(domain, 0)
            f.write(f"| {domain} | {score:.0f} | {cnt} |\n")

        f.write("\n### 情绪词出现频次\n\n")
        f.write("| 情绪 | 提及数 | 例句 |\n")
        f.write("|---|---|---|\n")
        for emotion, count in emotion_result["emotion_comment_counts"].items():
            examples = emotion_result["emotion_examples"].get(emotion, [])
            ex = examples[0][0][:50] + "…" if examples else ""
            f.write(f"| {emotion} | {count} | {ex} |\n")

        f.write("\n---\n\n")

        # 聚类摘要
        f.write("## 聚类摘要\n\n")
        cluster_method = clusters[0].get("method", "unknown") if clusters else "none"
        f.write(f"聚类方法：{cluster_method}（请求 {args.cluster} 类，实际 {len(clusters)} 类）\n\n")
        for cl in clusters:
            f.write(f"### 聚类 {cl['cluster_id']}: {cl['top_keywords'][:3]}（{cl['percentage']}%）\n")
            f.write(f"- **规模**: {cl['size']} 条 ({cl['percentage']}%)\n")
            f.write(f"- **高频词**: {'、'.join(cl['top_keywords'][:8])}\n")
            f.write(f"- **情绪基调**: {cl['emotional_tone']}\n")
            f.write("- **代表性评论**:\n")
            for q in cl["representative_quotes"][:3]:
                f.write(f"  - (👍{q['likes']}) {q['text']}\n")
            f.write("\n")

        f.write("---\n\n")

        # 所有评论导出（按点赞排序）
        f.write("## 全量评论（按互动排序 Top 100）\n\n")
        sorted_comments = sorted(comments, key=lambda x: -x["likes"])
        for i, c in enumerate(sorted_comments[:100], 1):
            loc = f"[{c['location']}] " if c.get("location") else ""
            f.write(f"{i}. ({loc}👍{c['likes']}) {c['text'][:150]}\n")

    print(f"✅ 初洗报告已生成: {report_path}")

    # ---- 5. 导出结构化 JSON（供后续 AI 读取） ----
    json_path = output_dir / "all_quotes.json"
    export_quotes = []
    for c in comments:
        export_quotes.append({
            "source_id": c["source_id"],
            "text": c["text"],
            "likes": c["likes"],
            "location": c.get("location", ""),
        })
    export_quotes.sort(key=lambda x: -x["likes"])

    cluster_by_source: dict[str, list[int]] = defaultdict(list)
    for cl in clusters:
        for source_id in cl.get("member_source_ids", []):
            cluster_by_source[source_id].append(cl["cluster_id"])
    for quote in export_quotes:
        quote["cluster_ids"] = cluster_by_source.get(quote["source_id"], [])

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({
            "total": total,
            "source": args.source,
            "quotes": export_quotes,
            "clusters": clusters,
            "freq_result": {
                "top_words": freq_result["top_words"],
                "top_nouns": freq_result.get("top_nouns", []),
                "top_adjs": freq_result.get("top_adjs", []),
                "top_verbs": freq_result.get("top_verbs", []),
            },
            "emotion": {
                "domain_scores": emotion_result["domain_scores"],
                "emotion_top": [
                    {"emotion": e, "count": c}
                    for e, c in emotion_result["emotion_comment_counts"].items()
                ],
            },
        }, f, ensure_ascii=False, indent=2)

    print(f"✅ 结构化数据已导出: {json_path}")
    print(f"\n📋 摘要:")
    print(f"   - 评论总数: {total}")
    print(f"   - 高频词数: {len(freq_result['top_words'])}")
    print(f"   - 聚类: {len(clusters)} 类")
    print(f"   - 情绪域: {', '.join(emotion_result['domain_scores'].keys())}")


if __name__ == "__main__":
    main()
